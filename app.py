from datetime import datetime
import multiprocessing
import sys
from tkinter.messagebox import *
from tkinter import scrolledtext, Tk, ttk, filedialog
from mttkinter import mtTkinter as mtk
import threading
import random
import time
from openpyxl import load_workbook
from emailtools import EmailTools


class EmailApplication:

    def __init__(self, master):
        self.root = master
        self.root.geometry("360x280")
        self.root.title("邮件自助工具 1.0")
        self.get_temp_info()
        self.__login_ui()

    def get_temp_info(self):
        wb = load_workbook("./template_info.xlsx")
        ws = wb.active
        self.temp_item = {}
        for line in list(ws.values)[1:]:
            self.temp_item[line[0]] = line[1]

    def __login_ui(self):
        self.login_box = mtk.LabelFrame(self.root, text="用户信息", fg="blue")
        self.login_box.place(x=20, y=20, width=320, height=200)

        user_email = mtk.Label(self.login_box, text="邮 箱：")
        user_email.place(x=20, y=20, width=80, height=30)
        self.email_text = mtk.Entry(self.login_box)
        self.email_text.place(x=100, y=20, width=170, height=30)
        user_pwd = mtk.Label(self.login_box, text="授权码：")
        user_pwd.place(x=20, y=80, width=80, height=30)
        self.pwd_text = mtk.Entry(self.login_box)
        self.pwd_text.place(x=100, y=80, width=170, height=30)
        self.home_btn = mtk.Button(self.login_box, text="进入主页", command=lambda: self.thread_it(self.__create_ui))
        self.home_btn.place(x=170, y=130, width=100, height=30)

    def __create_ui(self, *args):
        self.user_ = self.email_text.get()
        if not self.user_.strip():
            showerror("错误信息", "请输入邮箱账号.")
            return

        self.password_ = self.pwd_text.get()
        if not self.password_.strip():
            showerror("错误信息", "请输入授权码.")
            return
        home_win = mtk.Toplevel(self.root)
        home_win.title(f"邮件自助工具 1.0(当前账号：{self.user_})")
        home_win.geometry("560x380")
        self.msg_box = mtk.LabelFrame(home_win, text="统计信息", fg="blue")
        self.msg_box.place(x=20, y=20, width=250, height=80)

        self.msg_box_0 = mtk.Label(self.msg_box, text="当前无任何邮件发送任务.")
        self.msg_box_0.place(x=30, y=15, width=150, height=25)

        self.settings_box = mtk.LabelFrame(home_win, text="邮件发送设置", fg="blue")
        self.settings_box.place(x=20, y=120, width=250, height=130)
        # 模板选择
        email_temp = mtk.Label(self.settings_box, text="邮件模板：")
        email_temp.place(x=10, y=10, width=60, height=25)
        self.temp_num = ttk.Combobox(self.settings_box)
        self.temp_num["values"] = list(self.temp_item.keys()) + ["随机模板"]
        self.temp_num.place(x=80, y=10, width=150, height=25)
        # 随机间隔时间 10-30-50-60   设置开始时间
        random_sleep = mtk.Label(self.settings_box, text="间隔时间：")
        random_sleep.place(x=10, y=55, width=60, height=25)
        self.random_start = mtk.Entry(self.settings_box)
        self.random_start.place(x=80, y=55, width=50, height=25)
        random_ = mtk.Label(self.settings_box, text="~")
        random_.place(x=135, y=55, width=30, height=25)
        self.random_end = mtk.Entry(self.settings_box)
        self.random_end.place(x=170, y=55, width=50, height=25)
        # 提示信息框
        self.log_box = mtk.LabelFrame(home_win, text="提示信息", fg="blue")
        self.log_box.place(x=290, y=20, width=250, height=330)
        self.logtext = scrolledtext.ScrolledText(self.log_box, fg="green")
        self.logtext.place(x=15, y=5, width=230, height=290)

        # 任务开始栏
        self.task_box = mtk.LabelFrame(home_win)
        self.task_box.place(x=20, y=270, width=250, height=80)

        self.load_file_btn = mtk.Button(self.task_box, text="导入客户清单", command=lambda: self.thread_it(self.__load_excel))
        self.load_file_btn.place(x=10, y=20, width=120, height=40)

        self.task_start_btn = mtk.Button(self.task_box, text="开    始", command=lambda: self.thread_it(self.__start))
        self.task_start_btn.place(x=150, y=20, width=80, height=40)

    def __load_excel(self):
        excelPath = filedialog.askopenfilename(title=u'选择文件')
        if not excelPath.endswith(".xlsx"):
            showerror("错误信息", "Excel格式错误.")

        wb = load_workbook(excelPath)
        ws = wb.active
        self.email_tasks = [[str(i[0]), i[1], i[2]] for i in list(ws.values)[1:] if i]
        self.task_total = ws.max_row - 1
        if self.task_total <= 0:
            showerror("错误信息", f"未获取到客户Email信息.")
            return

        self.msg_box_0.config(text=f"当前客户Email数量：  0/{self.task_total}")

    def __start(self):
        if not self.email_tasks:
            showerror("错误信息", "当前无任何Email信息.")
            return

        # 间隔时间
        slp_start = self.random_start.get()
        if not slp_start:
            slp_start = 10

        slp_end = self.random_end.get()
        if not slp_end:
            slp_end = 30

        temp_num = self.temp_num.get()
        if not temp_num.strip():
            showerror("错误信息", "请选择Email模板.")
            return

        if temp_num == "随机模板":
            temp_num = random.choice(list(self.temp_item.values())[:-1])
        temp_title = self.temp_item.get(temp_num)
        email_sender = EmailTools(username=self.user_, password=self.password_)

        index = 1
        for email_, dis_, comp_exe, file in self.email_tasks:
            title = temp_title.replace("DistributorName", dis_)
            receiver = email_
            if file:
                email_sender.email_send(title=title, receiver=receiver, sender=self.user_, tem_name=temp_num,
                                        comp_exe=comp_exe, file=file)
            else:
                email_sender.email_send(title=title, receiver=receiver, sender=self.user_, tem_name=temp_num,
                                        comp_exe=comp_exe, file=file)

            self.msg_box_0.config(text=f"当前客户Email数量：  {index}/{self.task_total}")
            self.addLog(f"{email_}, 发送成功.")
            time.sleep(random.uniform(slp_start, slp_end))

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def addLog(self, msg):
        self.logtext.insert(mtk.END, "{} {}\n".format(datetime.now().strftime("%H:%M:%S"), msg))
        self.logtext.yview_moveto(1.0)


if __name__ == '__main__':
    if sys.platform.startswith('win'):
        multiprocessing.freeze_support()
    root = Tk()
    app = EmailApplication(root)
    root.mainloop()
